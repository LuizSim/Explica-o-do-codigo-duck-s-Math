from flask import Flask, request, jsonify, send_file  # Importa módulos do Flask para criar um servidor web, receber requisições HTTP, retornar respostas JSON e enviar arquivos
from flask_cors import CORS  # Importa a extensão Flask-CORS para permitir que o servidor aceite requisições de diferentes domínios, como de um navegador rodando um script JavaScript
import pandas as pd  # Importa a biblioteca pandas, que é usada principalmente para manipulação e análise de dados (neste código não foi utilizada, mas está pronta para ser usada)
import os  # Importa a biblioteca "os", que oferece funções para interagir com o sistema operacional, como manipular arquivos e diretórios
from openpyxl import Workbook  # Importa "Workbook" da biblioteca openpyxl para criar e manipular arquivos de planilhas Excel
from openpyxl.styles import PatternFill  # Importa "PatternFill" para aplicar cores a células específicas no Excel
from pathlib import Path  # Importa "Path" do módulo pathlib para lidar com caminhos de diretórios de maneira mais intuitiva e segura

# Cria uma instância da aplicação Flask, que será responsável por criar e gerenciar o servidor web
app = Flask(__name__)

# Permite que o servidor aceite requisições de diferentes origens, resolvendo problemas de segurança ao compartilhar dados entre diferentes domínios
CORS(app)

# Lista onde serão armazenadas as respostas enviadas pelo jogador durante o jogo
responses = []

# Cria uma rota na URL base '/' que pode receber requisições do tipo POST para salvar as respostas dos jogadores
@app.route('/', methods=['POST'])
def save_response():
    # Verifica se a requisição recebida é do tipo POST, que significa que o cliente está enviando dados para o servidor
    if request.method == 'POST':
        # Extrai os dados enviados no corpo da requisição, que devem estar no formato JSON, e armazena na variável "data"
        data = request.json
        # Adiciona a resposta recebida à lista "responses", que contém todas as respostas dos jogadores
        responses.append(data)
        # Retorna uma mensagem de confirmação dizendo que a resposta foi registrada com sucesso e o código HTTP 200 (sucesso)
        return jsonify({"message": "Resposta registrada com sucesso"}), 200

# Cria uma outra rota na URL base '/' que pode ser acessada por meio de requisições do tipo GET para gerar um arquivo Excel com as respostas dos jogadores
@app.route('/', methods=['GET'])
def generate_excel():
    # Verifica se a lista "responses" contém respostas registradas, ou seja, se houve participação dos jogadores
    if responses:
        # Define o caminho onde o arquivo Excel será salvo, neste caso na pasta de Downloads do usuário que está rodando o script
        caminho_pasta_downloads = str(Path.home() / "Downloads")

        # Define o nome do arquivo a ser gerado e combina com o caminho da pasta Downloads para formar o caminho completo do arquivo
        caminho_arquivo = os.path.join(caminho_pasta_downloads, 'resultado_ducks_math.xlsx')
        
        try:
            # Cria uma nova planilha (workbook) em branco usando a biblioteca openpyxl para armazenar as informações do jogo
            wb = Workbook()
            # A aba ativa do workbook é selecionada e seu título é alterado para "Respostas"
            ws = wb.active
            ws.title = "Respostas"

            # Cria o cabeçalho da planilha com os títulos das colunas, que explicam o conteúdo de cada coluna
            ws.append(["Perguntas", "Resposta", "Resposta do jogador", "Acertos", "Score"])

            # Define cores de preenchimento para as células de "Acertos": verde para respostas corretas e vermelho para respostas incorretas
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Itera sobre cada resposta registrada na lista "responses" e escreve seus dados na planilha
            for response in responses:
                # Extrai os detalhes da resposta, como a pergunta, a resposta correta, a resposta do jogador e a pontuação
                pergunta = response.get("question")
                resposta_correta = response.get("correct_answer")
                resposta_jogador = response.get("player_answer")
                # Verifica se a resposta do jogador é igual à resposta correta e define "Correto" ou "Incorreto"
                acerto = "Correto" if resposta_correta == resposta_jogador else "Incorreto"
                score = response.get("score", 0)

                # Adiciona todos esses valores na última linha da planilha, criando um novo registro para cada resposta
                ws.append([pergunta, resposta_correta, resposta_jogador, acerto, score])

                # Aplica a cor verde ou vermelha na célula da coluna "Acertos" dependendo se a resposta foi correta ou não
                acertos_cell = ws.cell(row=ws.max_row, column=4)
                if acerto == "Correto":
                    acertos_cell.fill = green_fill  # Colore a célula de verde se a resposta estiver correta
                else:
                    acertos_cell.fill = red_fill  # Colore a célula de vermelho se a resposta estiver incorreta

            # Salva a planilha na pasta Downloads do usuário com o nome "resultado_ducks_math.xlsx"
            wb.save(caminho_arquivo)

            # Retorna uma resposta indicando que o arquivo Excel foi gerado com sucesso, além de fornecer o caminho do arquivo gerado
            return jsonify({"message": "Arquivo Excel gerado com sucesso", "file_path": caminho_arquivo}), 200

        # Caso aconteça algum erro durante a criação e salvamento da planilha, captura o erro e retorna uma mensagem de erro ao cliente
        except Exception as e:
            return jsonify({"message": f"Erro ao gerar o arquivo Excel: {str(e)}"}), 500

    # Caso a lista "responses" esteja vazia, retorna uma mensagem informando que não há respostas registradas
    return jsonify({"message": "Nenhuma resposta registrada"}), 400

# Função principal que roda o servidor Flask, se o script for executado diretamente
if __name__ == '__main__':
    app.run(debug=True)  # Inicia o servidor Flask em modo de depuração, permitindo ver mensagens de erro no console

